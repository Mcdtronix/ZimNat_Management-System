# views.py - Updated with proper template integration
from rest_framework import generics, status, viewsets, filters, permissions
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.permissions import IsAuthenticated, AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth import login, authenticate, logout
from django.db.models import Count, Sum, Q
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json
from django.contrib import messages
from django.conf import settings
from django.utils.http import url_has_allowed_host_and_scheme
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from datetime import timedelta
from django.conf import settings
from django.http import JsonResponse
from django.http import HttpResponse
from rest_framework_simplejwt.tokens import RefreshToken
from django.db import transaction
from drf_spectacular.utils import extend_schema, inline_serializer, extend_schema_view
from rest_framework import serializers as rf_serializers

from .models import (
    User, Customer, Vehicle, VehicleCategory, InsuranceCoverage,
    InsurancePolicy, Claim, ClaimDocument, ContactInquiry,
    DashboardStats, Payment, Notification, Quotation, TwoFactorCode
)
from .serializers import (
    UserSerializer,
    RegisterSerializer,
    CustomerSerializer, VehicleSerializer, VehicleCategorySerializer,
    InsuranceCoverageSerializer, InsurancePolicySerializer, ClaimSerializer,
    ContactInquirySerializer, DashboardStatsSerializer, PaymentSerializer,
    QuoteResponseSerializer,
    NotificationSerializer,
    QuotationSerializer,
    ClaimDocumentSerializer,
)
from .utils import send_otp
from .permissions import IsOwnerOrReadOnly, IsStaffOrReadOnly, CustomerDataAccessPermission, IsCustomerOwnerOnly, CanProcessClaims, CanApproveClaims, IsManager, IsUnderwriter
from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.core.mail import send_mail
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth.password_validation import validate_password
from io import BytesIO

# Authentication Views (basic login/logout helpers using session or token can be added later if needed)

# Dashboard Views
class DashboardOverviewView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        if user.user_type == 'customer':
            return self.get_customer_dashboard()
        elif user.user_type in ['manager', 'underwriter']:
            return self.get_staff_dashboard()
        else:
            return self.get_general_dashboard()
    
    def get_customer_dashboard(self):
        try:
            customer = self.request.user.customer_profile
            policies = InsurancePolicy.objects.filter(customer=customer)
            claims = Claim.objects.filter(policy__customer=customer)
            
            data = {
                'active_policies': policies.filter(status='active').count(),
                'total_claims': claims.count(),
                'pending_claims': claims.filter(status__in=['submitted', 'under_review']).count(),
                'vehicles': customer.vehicles.count(),
                'recent_claims': ClaimSerializer(
                    claims.order_by('-created_at')[:5], many=True
                ).data
            }
            return Response(data)
        except:
            return Response({
                'active_policies': 0,
                'total_claims': 0,
                'pending_claims': 0,
                'vehicles': 0,
                'recent_claims': []
            })
    
    def get_staff_dashboard(self):
        # Chart data for the dashboard
        current_year = timezone.now().year
        chart_data = []
        
        for year in range(current_year - 4, current_year + 1):
            year_data = DashboardStats.objects.filter(year=year).first()
            if year_data:
                chart_data.append({
                    'year': year,
                    'value': year_data.claims_count,
                    'label': str(year)
                })
            else:
                chart_data.append({
                    'year': year,
                    'value': 0,
                    'label': str(year)
                })
        
        # Recent claims for approval
        recent_claims = Claim.objects.filter(
            approval_status='pending'
        ).select_related('policy__customer', 'policy__vehicle').order_by('-created_at')[:10]
        
        data = {
            'chart_data': chart_data,
            'recent_claims': ClaimSerializer(recent_claims, many=True).data,
            'total_customers': Customer.objects.count(),
            'active_policies': InsurancePolicy.objects.filter(status='active').count(),
            'pending_claims': Claim.objects.filter(approval_status='pending').count(),
        }
        return Response(data)
    
    def get_general_dashboard(self):
        data = {
            'total_customers': Customer.objects.count(),
            'active_policies': InsurancePolicy.objects.filter(status='active').count(),
            'total_claims': Claim.objects.count(),
            'pending_claims': Claim.objects.filter(status='pending').count(),
        }
        return Response(data)

# Customer Views
@extend_schema_view(list=extend_schema(tags=['customers']), retrieve=extend_schema(tags=['customers']), create=extend_schema(tags=['customers']), update=extend_schema(tags=['customers']), partial_update=extend_schema(tags=['customers']), destroy=extend_schema(tags=['customers']))
class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated, CustomerDataAccessPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['user__first_name', 'user__last_name', 'customer_id', 'user__email']
    filterset_fields = ['town', 'date_registered']
    
    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                return Customer.objects.filter(id=user.customer_profile.id)
            except AttributeError:
                return Customer.objects.none()
        return super().get_queryset()

class CustomerListView(generics.ListAPIView):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['user__first_name', 'user__last_name', 'customer_id']

# Vehicle Views
@extend_schema_view(list=extend_schema(tags=['vehicle-categories']), retrieve=extend_schema(tags=['vehicle-categories']))
class VehicleCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = VehicleCategory.objects.filter(is_active=True)
    serializer_class = VehicleCategorySerializer
    permission_classes = [AllowAny]

@extend_schema_view(list=extend_schema(tags=['vehicles']), retrieve=extend_schema(tags=['vehicles']), create=extend_schema(tags=['vehicles']), update=extend_schema(tags=['vehicles']), partial_update=extend_schema(tags=['vehicles']), destroy=extend_schema(tags=['vehicles']))
class VehicleViewSet(viewsets.ModelViewSet):
    queryset = Vehicle.objects.all()
    serializer_class = VehicleSerializer
    permission_classes = [IsAuthenticated, CustomerDataAccessPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['category', 'customer']

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                return Vehicle.objects.filter(customer=user.customer_profile)
            except Exception:
                return Vehicle.objects.none()
        return super().get_queryset()

    def perform_create(self, serializer):
        user = self.request.user
        if user.user_type != 'customer':
            return serializer.save()
        try:
            customer = user.customer_profile
        except Exception:
            raise rf_serializers.ValidationError({'customer': 'Customer profile not found.'})
        serializer.save(customer=customer)

    def perform_update(self, serializer):
        instance = self.get_object()
        user = self.request.user
        if user.user_type == 'customer':
            try:
                customer = user.customer_profile
            except Exception:
                raise rf_serializers.ValidationError({'customer': 'Customer profile not found.'})
            if instance.customer_id != customer.id:
                raise permissions.PermissionDenied('You can only modify your own vehicles.')
            serializer.save(customer=customer)
        else:
            serializer.save()

# Insurance Views
@extend_schema_view(list=extend_schema(tags=['insurance-coverages']), retrieve=extend_schema(tags=['insurance-coverages']))
class InsuranceCoverageViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = InsuranceCoverage.objects.filter(is_active=True)
    serializer_class = InsuranceCoverageSerializer
    permission_classes = [AllowAny]

@extend_schema_view(list=extend_schema(tags=['policies']), retrieve=extend_schema(tags=['policies']), create=extend_schema(tags=['policies']), update=extend_schema(tags=['policies']), partial_update=extend_schema(tags=['policies']), destroy=extend_schema(tags=['policies']))
class InsurancePolicyViewSet(viewsets.ModelViewSet):
    queryset = InsurancePolicy.objects.all()
    serializer_class = InsurancePolicySerializer
    permission_classes = [IsAuthenticated, CustomerDataAccessPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'customer', 'coverage']

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                return InsurancePolicy.objects.filter(customer=user.customer_profile)
            except:
                return InsurancePolicy.objects.none()
        return super().get_queryset()

    def perform_create(self, serializer):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                customer = user.customer_profile
            except Exception:
                raise rf_serializers.ValidationError({'customer': 'Customer profile not found.'})
            vehicle = serializer.validated_data.get('vehicle')
            if not vehicle or vehicle.customer_id != customer.id:
                raise permissions.PermissionDenied('You can only apply a policy for your own vehicle.')
            # Force pending on application and attach customer
            policy = serializer.save(customer=customer, status='pending')
        else:
            policy = serializer.save()

        # Notify all underwriters of new application
        underwriters = User.objects.filter(user_type='underwriter')
        notifs = [
            Notification(
                recipient=u,
                title='New policy application',
                message=f'Policy application {policy.policy_number} pending review.',
                type='status_update',
                payload={'policy_id': policy.id, 'policy_number': policy.policy_number, 'status': policy.status}
            ) for u in underwriters
        ]
        if notifs:
            Notification.objects.bulk_create(notifs)

    def perform_update(self, serializer):
        user = self.request.user
        if user.user_type == 'customer':
            raise permissions.PermissionDenied('Customers cannot modify policies after application.')
        serializer.save()

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def quote(self, request, pk=None):
        policy = self.get_object()
        if request.user.user_type != 'underwriter':
            raise permissions.PermissionDenied('Only underwriters can quote policies.')
        if policy.status not in ['pending']:
            return Response({'error': 'Only pending policies can be quoted.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            premium = Decimal(str(request.data.get('premium_amount')))
            coverage_amt = Decimal(str(request.data.get('coverage_amount')))
        except Exception:
            return Response({'error': 'premium_amount and coverage_amount must be decimals'}, status=status.HTTP_400_BAD_REQUEST)
        if premium <= 0 or coverage_amt <= 0:
            return Response({'error': 'premium_amount and coverage_amount must be positive.'}, status=status.HTTP_400_BAD_REQUEST)
        # Update monetary figures on policy for visibility
        policy.premium_amount = premium
        policy.coverage_amount = coverage_amt
        policy.save(update_fields=['premium_amount', 'coverage_amount'])

        # Defaults and payload
        default_terms = request.data.get('terms') or (
            'By accepting this quotation, you agree to the policy terms and conditions. '
            'Cover is activated upon receipt and confirmation of payment. '
            'Premiums are payable within 7 days of quotation unless otherwise agreed.'
        )
        bank_details = {
            'bank': 'CBZ',
            'account_number': '020224850175100',
            'branch': 'Chivhu',
        }
        payment_url = request.data.get('payment_url') or f"/payments/checkout?policy={policy.id}"

        # Create a Quotation record
        quote = Quotation.objects.create(
            policy=policy,
            premium_amount=premium,
            coverage_amount=coverage_amt,
            currency=request.data.get('currency', 'USD'),
            status='sent',
            terms=default_terms,
            bank_details=bank_details,
            payment_url=payment_url,
            created_by=request.user,
        )

        # Notify customer with quotation
        Notification.objects.create(
            recipient=policy.customer.user,
            title='Policy quotation available',
            message=f'Quotation for {policy.policy_number}',
            type='quotation',
            payload={
                'policy_id': policy.id,
                'policy_number': policy.policy_number,
                'premium_amount': str(premium),
                'coverage_amount': str(coverage_amt),
                'currency': quote.currency,
                'quote_id': quote.quote_id,
                'quotation_id': quote.id,
                'payment_url': payment_url,
                'bank_details': bank_details,
                'terms': default_terms,
            }
        )

        return Response({'message': 'Quotation created and customer notified.', 'quote_id': quote.quote_id})

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def auto_quote(self, request, pk=None):
        """
        Compute a subscription (quotation) from policy data using underwriting guideline
        and generate/send a quotation to the customer.

        Guideline (example for comprehensive):
        - Base premium: rate% of vehicle market value (default 6%)
        - Stamp duty: 5% of base
        - Motor levy: 2.25% of base
        - Annual premium = base + stamp + levy
        - Termly premium = annual / 3 (informational)
        """
        policy = self.get_object()
        # Permissions
        if request.user.user_type not in ['underwriter', 'manager']:
            raise permissions.PermissionDenied('Only underwriters or managers can auto-quote policies.')
        if policy.status != 'pending':
            return Response({'error': f'Only pending policies can be quoted. Current status: {policy.status}'}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure we have related objects
        policy.refresh_from_db()
        policy_vehicle = getattr(policy, 'vehicle', None)
        policy_coverage = getattr(policy, 'coverage', None)
        if not policy_vehicle or not policy_coverage:
            return Response({'error': 'Policy must have vehicle and coverage set.'}, status=status.HTTP_400_BAD_REQUEST)

        from decimal import Decimal, ROUND_HALF_UP
        market_value = policy_vehicle.market_value or Decimal('0')
        coverage_type = policy_coverage.name  # 'comprehensive' | 'third_party'

        # Default guideline rates (can be moved to settings)
        BASE_RATE = Decimal('0.06') if coverage_type == 'comprehensive' else Decimal('0.03')
        STAMP_DUTY_RATE = Decimal('0.05')
        MOTOR_LEVY_RATE = Decimal('0.0225') if coverage_type == 'comprehensive' else Decimal('0.015')

        base = (market_value * BASE_RATE).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        stamp = (base * STAMP_DUTY_RATE).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        levy = (base * MOTOR_LEVY_RATE).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        annual_premium = (base + stamp + levy).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        # Optional informational: termly = annual/3
        termly_premium = (annual_premium / Decimal('3')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        # If preview mode, DO NOT persist or notify
        preview = str(request.query_params.get('preview') or request.data.get('preview') or '').lower() in ['1', 'true', 'yes']
        if preview:
            default_terms = (
                'This quotation is based on underwriting guideline rates. Cover is activated upon payment. '
                f'Base: {BASE_RATE*100}% of sum insured; Stamp duty: {STAMP_DUTY_RATE*100}%; Motor levy: {MOTOR_LEVY_RATE*100}%. '
                f'Annual premium: USD {annual_premium}; Termly premium: USD {termly_premium}.'
            )
            return Response({
                'preview': True,
                'annual_premium': str(annual_premium),
                'termly_premium': str(termly_premium),
                'coverage_amount': str(market_value),
                'currency': 'USD',
                'terms': default_terms,
                'base_rate': str(BASE_RATE),
                'stamp_duty_rate': str(STAMP_DUTY_RATE),
                'motor_levy_rate': str(MOTOR_LEVY_RATE),
            })

        # Persist premium/coverage on policy for visibility
        policy.premium_amount = annual_premium
        policy.coverage_amount = market_value
        policy.save(update_fields=['premium_amount', 'coverage_amount'])

        # Create quotation record
        default_terms = (
            'This quotation is based on underwriting guideline rates. Cover is activated upon payment. '
            f'Base: {BASE_RATE*100}% of sum insured; Stamp duty: {STAMP_DUTY_RATE*100}%; Motor levy: {MOTOR_LEVY_RATE*100}%. '
            f'Annual premium: USD {annual_premium}; Termly premium: USD {termly_premium}.'
        )
        bank_details = {
            'bank': 'CBZ',
            'account_number': '020224850175100',
            'branch': 'Chivhu',
        }
        payment_url = f"/payments/checkout?policy={policy.id}"

        quote = Quotation.objects.create(
            policy=policy,
            premium_amount=annual_premium,
            coverage_amount=market_value,
            currency='USD',
            status='sent',
            terms=default_terms,
            bank_details=bank_details,
            payment_url=payment_url,
            created_by=request.user,
        )

        # Notify customer
        Notification.objects.create(
            recipient=policy.customer.user,
            title='Policy quotation available',
            message=f'Quotation for {policy.policy_number}',
            type='quotation',
            payload={
                'policy_id': policy.id,
                'policy_number': policy.policy_number,
                'premium_amount': str(annual_premium),
                'coverage_amount': str(market_value),
                'currency': quote.currency,
                'quote_id': quote.quote_id,
                'quotation_id': quote.id,
                'payment_url': payment_url,
                'bank_details': bank_details,
                'terms': default_terms,
            }
        )

        return Response({
            'message': 'Quotation computed and sent to customer.',
            'annual_premium': str(annual_premium),
            'termly_premium': str(termly_premium),
            'quote_id': quote.quote_id,
        })

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def message(self, request, pk=None):
        """Allow an underwriter/manager to manually send a message to the policy owner."""
        policy = self.get_object()
        if request.user.user_type not in ['underwriter', 'manager']:
            raise permissions.PermissionDenied('Only underwriters or managers can send messages to customers.')
        msg = (request.data.get('message') or '').strip()
        title = (request.data.get('title') or 'Message from Underwriter').strip()
        if not msg:
            return Response({'error': 'Message is required.'}, status=status.HTTP_400_BAD_REQUEST)
        Notification.objects.create(
            recipient=policy.customer.user,
            title=title,
            message=msg,
            type='message',
            payload={'policy_id': policy.id, 'policy_number': policy.policy_number}
        )
        return Response({'message': 'Message sent to customer.'})

    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def export_quote(self, request, pk=None):
        """
        Export a quotation as PDF or Excel. Works for preview or persisted quotes.
        Query params:
          - format: 'pdf' | 'xlsx'
          - preview: true|false (if true, compute current preview without persisting)
        """
        policy = self.get_object()
        fmt = (request.query_params.get('format') or 'pdf').lower()
        preview = str(request.query_params.get('preview') or 'false').lower() in ['1', 'true', 'yes']

        # Compute values
        from decimal import Decimal, ROUND_HALF_UP
        policy_vehicle = getattr(policy, 'vehicle', None)
        policy_coverage = getattr(policy, 'coverage', None)
        if not policy_vehicle or not policy_coverage:
            return Response({'error': 'Policy must have vehicle and coverage set.'}, status=status.HTTP_400_BAD_REQUEST)
        market_value = policy_vehicle.market_value or Decimal('0')
        coverage_type = policy_coverage.name
        BASE_RATE = Decimal('0.06') if coverage_type == 'comprehensive' else Decimal('0.03')
        STAMP_DUTY_RATE = Decimal('0.05')
        MOTOR_LEVY_RATE = Decimal('0.0225') if coverage_type == 'comprehensive' else Decimal('0.015')
        base = (market_value * BASE_RATE).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        stamp = (base * STAMP_DUTY_RATE).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        levy = (base * MOTOR_LEVY_RATE).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        annual_premium = (base + stamp + levy).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        termly_premium = (annual_premium / Decimal('3')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        title = 'Motor Comprehensive I-USD Quotation' if coverage_type == 'comprehensive' else 'Third Party Quotation'
        insured_name = policy.customer.user.get_full_name() or policy.customer.user.username
        vehicle_label = f"{policy_vehicle.make} {policy_vehicle.model}"
        reg_no = getattr(policy_vehicle, 'registration_number', '') or getattr(policy_vehicle, 'reg_no', '') or ''
        year = getattr(policy_vehicle, 'year', '') or ''

        if fmt == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
                from reportlab.lib.units import mm
                from reportlab.lib import colors
            except Exception:
                return Response({'error': "PDF export requires 'reportlab'"}, status=status.HTTP_501_NOT_IMPLEMENTED)
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4
            y = height - 30*mm
            # Header
            c.setFont('Helvetica-Bold', 16)
            c.drawString(20*mm, y, 'QUOTATION')
            y -= 8*mm
            c.setFont('Helvetica', 10)
            c.drawString(20*mm, y, f"Name of Insured: {insured_name or 'TBA'}")
            y -= 6*mm
            c.drawString(20*mm, y, f"Business Description: {title}")
            y -= 10*mm
            # Vehicle table (simple)
            c.setFont('Helvetica-Bold', 11)
            c.drawString(20*mm, y, 'Motor Vehicles')
            y -= 6*mm
            c.setFont('Helvetica', 10)
            c.drawString(22*mm, y, f"Make and model: {vehicle_label}   Year: {year}   Reg No: {reg_no}   Value: USD {market_value}")
            y -= 10*mm
            # Premium breakdown
            c.setFont('Helvetica-Bold', 11)
            c.drawString(20*mm, y, 'Premium Breakdown')
            y -= 6*mm
            c.setFont('Helvetica', 10)
            c.drawString(22*mm, y, f"Base Premium ({(BASE_RATE*100)}%): USD {base}")
            y -= 5*mm
            c.drawString(22*mm, y, f"Stamp Duty (5% of base): USD {stamp}")
            y -= 5*mm
            c.drawString(22*mm, y, f"Motor Levy ({(MOTOR_LEVY_RATE*100)}% of base): USD {levy}")
            y -= 6*mm
            c.setFont('Helvetica-Bold', 11)
            c.drawString(22*mm, y, f"TOTAL Annual Premium: USD {annual_premium}")
            y -= 5*mm
            c.setFont('Helvetica', 10)
            c.drawString(22*mm, y, f"Termly Premium (1/3): USD {termly_premium}")
            y -= 12*mm
            # Limits and excesses (static text per template)
            c.setFont('Helvetica-Bold', 11)
            c.drawString(20*mm, y, 'Own Damage Limits (summary)')
            y -= 6*mm
            c.setFont('Helvetica', 9)
            c.drawString(22*mm, y, 'Medical Expenses: USD 500.00   Towing: Reasonable to nearest garage   Emergency: USD 200.00')
            y -= 12*mm
            c.setFont('Helvetica', 8)
            c.setFillColor(colors.grey)
            c.drawString(20*mm, y, 'This document is a system-generated quotation preview. Cover is activated upon payment.')
            c.setFillColor(colors.black)
            c.showPage()
            c.save()
            buffer.seek(0)
            resp = HttpResponse(buffer.read(), content_type='application/pdf')
            filename = f"quotation_{policy.policy_number or policy.id}.pdf"
            resp['Content-Disposition'] = f"attachment; filename=\"{filename}\""
            return resp
        elif fmt in ['xlsx', 'xls']:
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
            except Exception:
                return Response({'error': "Excel export requires 'openpyxl'"}, status=status.HTTP_501_NOT_IMPLEMENTED)
            wb = Workbook()
            ws = wb.active
            ws.title = 'Quotation'
            rows = [
                ['QUOTATION'],
                ['Name of Insured', insured_name or 'TBA'],
                ['Business Description', title],
                [],
                ['Motor Vehicles'],
                ['Make and model', vehicle_label, 'Year', year, 'Reg No', reg_no, 'Value (USD)', float(market_value)],
                [],
                ['Premium Breakdown'],
                ['Base Premium', float(base)],
                ['Stamp Duty', float(stamp)],
                ['Motor Levy', float(levy)],
                ['TOTAL Annual Premium', float(annual_premium)],
                ['Termly Premium', float(termly_premium)],
            ]
            for r in rows:
                ws.append(r)
            # Autosize
            for col in ws.columns:
                max_length = 12
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            resp = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"quotation_{policy.policy_number or policy.id}.xlsx"
            resp['Content-Disposition'] = f"attachment; filename=\"{filename}\""
            return resp
        else:
            return Response({'error': 'Unsupported format. Use pdf or xlsx.'}, status=status.HTTP_400_BAD_REQUEST)


@extend_schema_view(list=extend_schema(tags=['quotations']), retrieve=extend_schema(tags=['quotations']), create=extend_schema(tags=['quotations']))
class QuotationViewSet(viewsets.ModelViewSet):
    queryset = Quotation.objects.all()
    serializer_class = QuotationSerializer
    permission_classes = [IsAuthenticated, CustomerDataAccessPermission]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['status', 'policy', 'policy__customer']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                return Quotation.objects.filter(policy__customer=user.customer_profile)
            except Exception:
                return Quotation.objects.none()
        return super().get_queryset()

    def perform_create(self, serializer):
        user = self.request.user
        if user.user_type not in ['underwriter', 'manager']:
            raise permissions.PermissionDenied('Only underwriters or managers can create quotations.')
        policy = serializer.validated_data.get('policy')
        if policy.status != 'pending':
            raise rf_serializers.ValidationError({'policy': 'Only pending policies can be quoted.'})
        quote = serializer.save(created_by=user, status='sent')
        # Notify customer
        Notification.objects.create(
            recipient=policy.customer.user,
            title='Policy quotation available',
            message=f'Quotation for {policy.policy_number}',
            type='quotation',
            payload={
                'policy_id': policy.id,
                'policy_number': policy.policy_number,
                'premium_amount': str(quote.premium_amount),
                'coverage_amount': str(quote.coverage_amount),
                'currency': quote.currency,
                'quote_id': quote.quote_id,
                'quotation_id': quote.id,
                'payment_url': quote.payment_url,
                'bank_details': quote.bank_details,
                'terms': quote.terms,
            }
        )

    @action(detail=True, methods=['post'])
    def accept(self, request, pk=None):
        quote = self.get_object()
        user = request.user
        # Only the owning customer can accept/decline
        if not quote.policy or quote.policy.customer.user_id != user.id:
            raise permissions.PermissionDenied('Only the policy owner can accept this quotation.')
        if quote.status != 'sent':
            return Response({'error': f'Only sent quotations can be accepted. Current status: {quote.status}'}, status=status.HTTP_400_BAD_REQUEST)
        quote.status = 'accepted'
        quote.customer_decision_at = timezone.now()
        quote.decided_by = user
        quote.save(update_fields=['status', 'customer_decision_at', 'decided_by'])
        # Notify underwriters
        for uw in User.objects.filter(user_type='underwriter'):
            Notification.objects.create(
                recipient=uw,
                title='Quotation accepted',
                message=f'{user.get_full_name() or user.username} accepted quote {quote.quote_id} for {quote.policy.policy_number}',
                type='status_update',
                payload={'quote_id': quote.quote_id, 'policy_id': quote.policy_id, 'status': 'accepted'}
            )
        return Response({'message': 'Quotation accepted. Proceed to payment.', 'payment_url': quote.payment_url})

    @action(detail=True, methods=['post'])
    def decline(self, request, pk=None):
        quote = self.get_object()
        user = request.user
        if not quote.policy or quote.policy.customer.user_id != user.id:
            raise permissions.PermissionDenied('Only the policy owner can decline this quotation.')
        if quote.status != 'sent':
            return Response({'error': f'Only sent quotations can be declined. Current status: {quote.status}'}, status=status.HTTP_400_BAD_REQUEST)
        quote.status = 'declined'
        quote.customer_decision_at = timezone.now()
        quote.decided_by = user
        quote.save(update_fields=['status', 'customer_decision_at', 'decided_by'])
        # Notify underwriters
        for uw in User.objects.filter(user_type='underwriter'):
            Notification.objects.create(
                recipient=uw,
                title='Quotation declined',
                message=f'{user.get_full_name() or user.username} declined quote {quote.quote_id} for {quote.policy.policy_number}',
                type='status_update',
                payload={'quote_id': quote.quote_id, 'policy_id': quote.policy_id, 'status': 'declined'}
            )
        return Response({'message': 'Quotation declined.'})

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def approve(self, request, pk=None):
        policy = self.get_object()
        if request.user.user_type != 'underwriter':
            raise permissions.PermissionDenied('Only underwriters can approve policies.')
        if policy.status != 'pending':
            return Response({'error': f'Only pending policies can be approved. Current status: {policy.status}'}, status=status.HTTP_400_BAD_REQUEST)
        if not policy.premium_amount or not policy.coverage_amount or policy.premium_amount <= 0 or policy.coverage_amount <= 0:
            return Response({'error': 'Policy must have a positive quoted premium and coverage before approval.'}, status=status.HTTP_400_BAD_REQUEST)
        policy.status = 'active'
        policy.save()
        Notification.objects.create(
            recipient=policy.customer.user,
            title='Policy approved',
            message=f'Your policy {policy.policy_number} has been approved and activated.',
            type='status_update',
            payload={'policy_id': policy.id, 'status': policy.status}
        )
        return Response({'message': 'Policy approved and customer notified.'})

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def reject(self, request, pk=None):
        policy = self.get_object()
        if request.user.user_type != 'underwriter':
            raise permissions.PermissionDenied('Only underwriters can reject policies.')
        if policy.status != 'pending':
            return Response({'error': f'Only pending policies can be rejected. Current status: {policy.status}'}, status=status.HTTP_400_BAD_REQUEST)
        policy.status = 'cancelled'
        policy.save()
        Notification.objects.create(
            recipient=policy.customer.user,
            title='Policy rejected',
            message=f'Your policy {policy.policy_number} has been rejected.',
            type='status_update',
            payload={'policy_id': policy.id, 'status': policy.status}
        )
        return Response({'message': 'Policy rejected and customer notified.'})

# Payment Views
@extend_schema_view(list=extend_schema(tags=['payments']), retrieve=extend_schema(tags=['payments']), create=extend_schema(tags=['payments']), update=extend_schema(tags=['payments']), partial_update=extend_schema(tags=['payments']), destroy=extend_schema(tags=['payments']))
class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [IsAuthenticated, CustomerDataAccessPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'payment_method', 'policy__customer']

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                return Payment.objects.filter(policy__customer=user.customer_profile)
            except:
                return Payment.objects.none()
        return super().get_queryset()

# Contact/Inquiry Views
@extend_schema_view(list=extend_schema(tags=['contact-inquiries']), retrieve=extend_schema(tags=['contact-inquiries']), create=extend_schema(tags=['contact-inquiries']), update=extend_schema(tags=['contact-inquiries']), partial_update=extend_schema(tags=['contact-inquiries']), destroy=extend_schema(tags=['contact-inquiries']))
class ContactInquiryViewSet(viewsets.ModelViewSet):
    queryset = ContactInquiry.objects.all()
    serializer_class = ContactInquirySerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['status']
    ordering = ['-created_at']

    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def respond(self, request, pk=None):
        inquiry = self.get_object()
        response_text = request.data.get('response', '')
        
        if not response_text:
            return Response(
                {'error': 'Response text is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        inquiry.response = response_text
        inquiry.responded_by = request.user
        inquiry.status = 'resolved'
        inquiry.save()
        
        return Response({
            'message': 'Response sent successfully',
            'inquiry_id': inquiry.id
        })

# Claims Views
@extend_schema_view(list=extend_schema(tags=['claims']), retrieve=extend_schema(tags=['claims']), create=extend_schema(tags=['claims']), update=extend_schema(tags=['claims']), partial_update=extend_schema(tags=['claims']), destroy=extend_schema(tags=['claims']))
class ClaimViewSet(viewsets.ModelViewSet):
    queryset = Claim.objects.all()
    serializer_class = ClaimSerializer
    permission_classes = [IsAuthenticated, CustomerDataAccessPermission]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['status', 'approval_status', 'policy__customer']
    ordering_fields = ['created_at', 'estimated_amount']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                return Claim.objects.filter(policy__customer=user.customer_profile).select_related('policy__vehicle', 'policy__customer')
            except:
                return Claim.objects.none()
        return super().get_queryset().select_related('policy__vehicle', 'policy__customer')

    def perform_create(self, serializer):
        user = self.request.user
        try:
            customer = user.customer_profile
        except Exception:
            raise rf_serializers.ValidationError({'policy': 'Customer profile not found.'})
        policy = serializer.validated_data.get('policy')
        if not policy or policy.customer_id != customer.id:
            raise permissions.PermissionDenied('You can only claim on your own policy.')
        # Ensure comprehensive only
        if policy.coverage.name != 'comprehensive':
            raise rf_serializers.ValidationError({'policy': 'Only comprehensive policies are eligible for claims.'})
        claim = serializer.save(status='submitted', approval_status='pending')

        # Notify all underwriters of new claim submission
        underwriters = User.objects.filter(user_type='underwriter')
        notifs = [
            Notification(
                recipient=u,
                title='New claim submitted',
                message=f'Claim {claim.claim_id} requires review.',
                type='status_update',
                payload={'claim_id': claim.id, 'claim_ref': claim.claim_id, 'policy_number': claim.policy.policy_number}
            ) for u in underwriters
        ]
        if notifs:
            Notification.objects.bulk_create(notifs)

    @action(detail=True, methods=['post'])
    def process_claim(self, request, pk=None):
        claim = self.get_object()
        action = request.data.get('action')
        approved_amount = request.data.get('approved_amount')
        # Only underwriters/managers can process
        if request.user.user_type not in ['underwriter', 'manager']:
            raise permissions.PermissionDenied('Only underwriters or managers can process claims.')
        # Only pending approvals can be processed
        if claim.approval_status != 'pending':
            return Response({'success': False, 'error': 'Only pending claims can be processed.'}, status=status.HTTP_400_BAD_REQUEST)

        if action not in ['approve', 'reject']:
            return Response({'success': False, 'error': 'Invalid action'}, status=status.HTTP_400_BAD_REQUEST)

        if action == 'approve':
            claim.approval_status = 'approve'
            claim.status = 'approved'
            if approved_amount is not None:
                try:
                    claim.approved_amount = Decimal(str(approved_amount))
                except Exception:
                    return Response({'success': False, 'error': 'approved_amount must be a decimal'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            claim.approval_status = 'reject'
            claim.status = 'rejected'

        claim.processed_by = request.user
        claim.save()

        # Notify customer on status change
        Notification.objects.create(
            recipient=claim.policy.customer.user,
            title='Claim status updated',
            message=f'Your claim {claim.claim_id} is now {claim.status}.',
            type='status_update',
            payload={'claim_id': claim.id, 'claim_ref': claim.claim_id, 'status': claim.status}
        )

        return Response({
            'success': True,
            'message': f'Claim {action}d successfully',
            'claim_id': claim.claim_id,
            'status': claim.approval_status
        })

    @action(detail=True, methods=['post'])
    def assessor_message(self, request, pk=None):
        """Underwriter/Manager sends assessor visit details to claimant as a notification."""
        claim = self.get_object()
        if request.user.user_type not in ['underwriter', 'manager']:
            raise permissions.PermissionDenied('Only underwriters or managers can send assessor messages.')

        assessor_name = request.data.get('assessor_name')
        assessor_phone = request.data.get('assessor_phone')
        visit_date = request.data.get('visit_date')  # Expect ISO date string
        extra_message = request.data.get('message', '')

        if not assessor_name or not visit_date:
            return Response({'success': False, 'error': 'assessor_name and visit_date are required'}, status=status.HTTP_400_BAD_REQUEST)

        # Create notification to the claimant (policy owner)
        Notification.objects.create(
            recipient=claim.policy.customer.user,
            title='Assessor Visit Scheduled',
            message=f'An assessor will visit on {visit_date}. {extra_message}'.strip(),
            type='message',
            payload={
                'claim_id': claim.id,
                'claim_ref': claim.claim_id,
                'policy_number': claim.policy.policy_number,
                'assessor_name': assessor_name,
                'assessor_phone': assessor_phone,
                'visit_date': visit_date,
            }
        )

        return Response({'success': True, 'message': 'Assessor message sent to claimant.'})

@extend_schema_view(list=extend_schema(tags=['claim-documents']), retrieve=extend_schema(tags=['claim-documents']), create=extend_schema(tags=['claim-documents']), destroy=extend_schema(tags=['claim-documents']))
class ClaimDocumentViewSet(viewsets.ModelViewSet):
    queryset = ClaimDocument.objects.all()
    serializer_class = ClaimDocumentSerializer
    permission_classes = [IsAuthenticated, CustomerDataAccessPermission]
    http_method_names = ['get', 'post', 'delete']

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                customer = user.customer_profile
            except Exception:
                return ClaimDocument.objects.none()
            return ClaimDocument.objects.filter(claim__policy__customer=customer)
        return super().get_queryset()

    def perform_create(self, serializer):
        claim = serializer.validated_data.get('claim')
        user = self.request.user
        if user.user_type == 'customer':
            try:
                customer = user.customer_profile
            except Exception:
                raise rf_serializers.ValidationError({'claim': 'Customer profile not found.'})
            if not claim or claim.policy.customer_id != customer.id:
                raise permissions.PermissionDenied('You can only upload documents for your own claims.')
        doc = serializer.save()

        # If required documents are now present, notify all underwriters with document links
        try:
            required_types = {'police_report', 'id_document'}
            existing_types = set(
                ClaimDocument.objects.filter(claim=claim).values_list('document_type', flat=True)
            )
            if required_types.issubset(existing_types):
                underwriters = User.objects.filter(user_type='underwriter')
                # Build absolute URLs for documents
                docs = []
                for d in ClaimDocument.objects.filter(claim=claim).order_by('uploaded_at'):
                    try:
                        url = self.request.build_absolute_uri(d.document.url)
                    except Exception:
                        url = d.document.url if hasattr(d.document, 'url') else ''
                    docs.append({
                        'id': d.id,
                        'type': d.document_type,
                        'url': url,
                        'uploaded_at': d.uploaded_at.isoformat() if hasattr(d.uploaded_at, 'isoformat') else str(d.uploaded_at),
                    })
                notifs = [
                    Notification(
                        recipient=u,
                        title='Claim documents submitted',
                        message=f'Claim {claim.claim_id} has all required documents attached and is ready for review.',
                        type='message',
                        payload={
                            'claim_id': claim.id,
                            'claim_ref': claim.claim_id,
                            'policy_number': claim.policy.policy_number,
                            'documents': docs,
                        }
                    ) for u in underwriters
                ]
                if notifs:
                    Notification.objects.bulk_create(notifs)
        except Exception:
            # Avoid failing the upload if notification assembly fails
            pass

# Notifications
@extend_schema_view(list=extend_schema(tags=['notifications']), retrieve=extend_schema(tags=['notifications']), partial_update=extend_schema(tags=['notifications']))
class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-created_at']

    def get_queryset(self):
        return Notification.objects.filter(recipient=self.request.user).order_by('-created_at')

    @action(detail=True, methods=['post'])
    def read(self, request, pk=None):
        notif = self.get_object()
        if notif.recipient_id != request.user.id:
            raise permissions.PermissionDenied('Not your notification.')
        notif.is_read = True
        notif.save(update_fields=['is_read'])
        return Response({'message': 'Notification marked as read.'})

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        """Return count of unread notifications for the authenticated user."""
        count = Notification.objects.filter(recipient=request.user, is_read=False).count()
        return Response({'count': count})

# Simple claims list for dashboards using the main serializer
class ClaimListView(generics.ListAPIView):
    serializer_class = ClaimSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['approval_status', 'status', 'policy__customer']

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'customer':
            try:
                return Claim.objects.filter(policy__customer=user.customer_profile)
            except:
                return Claim.objects.none()
        return Claim.objects.all()

# Dashboard Stats Management
@extend_schema_view(list=extend_schema(tags=['dashboard-stats']), retrieve=extend_schema(tags=['dashboard-stats']), create=extend_schema(tags=['dashboard-stats']), update=extend_schema(tags=['dashboard-stats']), partial_update=extend_schema(tags=['dashboard-stats']), destroy=extend_schema(tags=['dashboard-stats']))
class DashboardStatsViewSet(viewsets.ModelViewSet):
    queryset = DashboardStats.objects.all()
    serializer_class = DashboardStatsSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'])
    def update_stats(self, request):
        """Update dashboard statistics for all years"""
        current_year = timezone.now().year
        
        for year in range(current_year - 4, current_year + 1):
            year_start = datetime(year, 1, 1)
            year_end = datetime(year, 12, 31, 23, 59, 59)
            
            # Calculate statistics for the year
            claims_count = Claim.objects.filter(
                created_at__range=[year_start, year_end]
            ).count()
            
            policies_count = InsurancePolicy.objects.filter(
                created_at__range=[year_start, year_end]
            ).count()
            
            customers_count = Customer.objects.filter(
                date_registered__range=[year_start, year_end]
            ).count()
            
            revenue = InsurancePolicy.objects.filter(
                created_at__range=[year_start, year_end]
            ).aggregate(Sum('premium_amount'))['premium_amount__sum'] or 0
            
            # Update or create stats
            stats, created = DashboardStats.objects.update_or_create(
                year=year,
                defaults={
                    'claims_count': claims_count,
                    'policies_count': policies_count,
                    'customers_count': customers_count,
                    'revenue': revenue,
                }
            )
        
        return Response({'message': 'Dashboard statistics updated successfully'})

# User Profile Views
class UserProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    
    def get_object(self):
        return self.request.user

class CustomerProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]
    
    def get_object(self):
        try:
            return self.request.user.customer_profile
        except:
            return None

# Search Views
@extend_schema(tags=['search'], summary='Search customers')
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_customers(request):
    query = request.GET.get('q', '')
    if len(query) < 2:
        return Response({'results': []})
    
    customers = Customer.objects.filter(
        Q(user__first_name__icontains=query) |
        Q(user__last_name__icontains=query) |
        Q(customer_id__icontains=query) |
        Q(user__email__icontains=query)
    )[:10]
    
    serializer = CustomerSerializer(customers, many=True)
    return Response({'results': serializer.data})

@extend_schema(tags=['search'], summary='Search vehicles')
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_vehicles(request):
    query = request.GET.get('q', '')
    if len(query) < 2:
        return Response({'results': []})
    
    vehicles = Vehicle.objects.filter(
        Q(vehicle_number__icontains=query) |
        Q(make__icontains=query) |
        Q(model__icontains=query)
    )[:10]
    
    serializer = VehicleSerializer(vehicles, many=True)
    return Response({'results': serializer.data})

# Statistics and Analytics Views
@extend_schema(tags=['analytics'], summary='Analytics overview for charts and reports')
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def analytics_overview(request):
    """Provide analytics data for charts and reports.
    If the user is a customer, scope metrics to their own policies and claims only.
    """
    user = request.user
    # Monthly data for current year
    current_year = timezone.now().year
    monthly_data = []

    # Determine scoping filters
    customer_filter = {}
    policy_customer_filter = {}
    if getattr(user, 'user_type', None) == 'customer':
        try:
            customer = user.customer_profile
            customer_filter = { 'policy__customer': customer }
            policy_customer_filter = { 'customer': customer }
        except Exception:
            # No customer profile; keep filters empty which will yield 0s with explicit guards
            customer_filter = { 'policy__customer': None }
            policy_customer_filter = { 'customer': None }

    for month in range(1, 13):
        month_start = datetime(current_year, month, 1)
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = datetime(current_year, month + 1, 1) - timedelta(days=1)

        claims_qs = Claim.objects.filter(created_at__range=[month_start, month_end])
        policies_qs = InsurancePolicy.objects.filter(created_at__range=[month_start, month_end])

        if customer_filter:
            claims_qs = claims_qs.filter(**customer_filter)
        if policy_customer_filter:
            policies_qs = policies_qs.filter(**policy_customer_filter)

        monthly_data.append({
            'month': month,
            'month_name': month_start.strftime('%B'),
            'claims': claims_qs.count(),
            'policies': policies_qs.count(),
        })

    # Category distribution (non-sensitive; for customers we can scope to their vehicles)
    category_data = []
    categories = VehicleCategory.objects.all()
    for category in categories:
        vehicle_qs = Vehicle.objects.filter(category=category)
        if policy_customer_filter:
            vehicle_qs = vehicle_qs.filter(customer=policy_customer_filter.get('customer'))
        count = vehicle_qs.count()
        category_data.append({
            'category': category.get_name_display(),
            'count': count
        })

    return Response({
        'monthly_data': monthly_data,
        'category_distribution': category_data,
        'year': current_year
    })

# Quote Generation Views
@extend_schema(tags=['quotes'], summary='Generate a quick insurance quote')
@api_view(['POST'])
@permission_classes([AllowAny])
def generate_quote(request):
    try:
        market_value = Decimal(str(request.data.get('market_value')))
        coverage_type = request.data.get('coverage_type')
        vehicle_category = request.data.get('vehicle_category')
    except Exception:
        return Response({'error': 'Invalid input'}, status=status.HTTP_400_BAD_REQUEST)

    if coverage_type not in ['third_party', 'comprehensive']:
        return Response({'error': 'coverage_type must be third_party or comprehensive'}, status=status.HTTP_400_BAD_REQUEST)

    base_rate = Decimal('0.05') if coverage_type == 'third_party' else Decimal('0.08')
    category_multipliers = {
        'motorcycles': Decimal('1.2'),
        'light_motor': Decimal('1.0'),
        'minibuses': Decimal('1.3'),
        'buses': Decimal('1.4'),
        'heavy_vehicles': Decimal('1.6'),
        'haulage_trucks': Decimal('1.8'),
    }
    multiplier = category_multipliers.get(vehicle_category, Decimal('1.0'))
    estimated_premium = market_value * base_rate * multiplier

    if coverage_type == 'third_party':
        coverage_amount = min(market_value, Decimal('1000000'))
        coverage_details = [
            'Third party injury coverage',
            'Third party property damage up to limit',
            'Does not cover your vehicle',
        ]
    else:
        coverage_amount = market_value
        coverage_details = [
            'Comprehensive vehicle coverage',
            'Third party coverage',
            'Own damage coverage',
            'Theft and fire protection',
            'Natural disaster coverage',
        ]

    data = {
        'estimated_premium': estimated_premium,
        'coverage_amount': coverage_amount,
        'coverage_details': coverage_details,
        'validity_period': 30,
    }
    return Response(data)

# Report Generation Views
@extend_schema(tags=['reports'], summary='Generate summary reports')
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_report(request):
    report_type = request.data.get('report_type', 'claims')
    start_date = request.data.get('start_date')
    end_date = request.data.get('end_date')

    queryset = Claim.objects.all()
    if start_date:
        queryset = queryset.filter(created_at__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(created_at__date__lte=end_date)

    data = {
        'total_claims': queryset.count(),
        'approved_claims': queryset.filter(approval_status='approve').count(),
        'rejected_claims': queryset.filter(approval_status='reject').count(),
        'pending_claims': queryset.filter(approval_status='pending').count(),
        'total_amount': queryset.aggregate(Sum('estimated_amount'))['estimated_amount__sum'] or 0,
        'approved_amount': queryset.filter(approval_status='approve').aggregate(
            Sum('approved_amount'))['approved_amount__sum'] or 0,
    }

    return Response({
        'report_type': report_type,
        'period': {
            'start_date': start_date,
            'end_date': end_date,
        },
        'data': data,
        'generated_at': timezone.now(),
    })

# Utility Views
@extend_schema(tags=['utility'], summary='Health check')
@api_view(['GET'])
@permission_classes([AllowAny])
def health_check(request):
    return Response({
        'status': 'healthy',
        'timestamp': timezone.now(),
        'version': '1.0.0'
    })

@extend_schema(tags=['auth'], summary='Get current user permissions')
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_permissions(request):
    """Return user permissions and profile data for frontend routing."""
    user = request.user
    
    # Base user data
    user_data = {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'user_type': user.user_type,
        'is_active': user.is_active,
    }
    
    # Role-specific permissions and data
    permissions = {
        'can_view_dashboard': True,
        'can_manage_vehicles': user.user_type in ['customer', 'manager', 'underwriter'],
        'can_apply_policies': user.user_type == 'customer',
        'can_submit_claims': user.user_type == 'customer',
        'can_process_claims': user.user_type in ['manager', 'underwriter'],
        'can_approve_claims': user.user_type == 'manager',
        'can_quote_policies': user.user_type in ['manager', 'underwriter'],
        'can_manage_customers': user.user_type in ['manager', 'underwriter'],
        'can_view_all_data': user.user_type in ['manager', 'underwriter'],
        'is_staff': user.user_type in ['manager', 'underwriter'] or user.is_staff,
        'is_customer': user.user_type == 'customer',
        'is_manager': user.user_type == 'manager',
        'is_underwriter': user.user_type == 'underwriter',
    }
    
    # Customer-specific data
    customer_data = None
    if user.user_type == 'customer':
        try:
            customer = user.customer_profile
            customer_data = {
                'customer_id': customer.customer_id,
                'address_no': customer.address_no,
                'street': customer.street,
                'town': customer.town,
                'date_registered': customer.date_registered,
            }
        except AttributeError:
            # Customer profile doesn't exist
            permissions['needs_profile_setup'] = True
    
    return Response({
        'user': user_data,
        'permissions': permissions,
        'customer_profile': customer_data,
        'dashboard_route': _get_dashboard_route(user.user_type),
    })

def _get_dashboard_route(user_type):
    """Return appropriate dashboard route based on user type."""
    if user_type == 'customer':
        return '/dashboard'
    elif user_type in ['manager', 'underwriter']:
        return '/admin-dashboard'
    else:
        return '/dashboard'

# Template Views
class HomeTemplateView(TemplateView):
    template_name = 'home.html'

def home_view(request):
    return render(request, 'home.html')

def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('home')

def login_page_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    next_url = request.GET.get('next') or request.POST.get('next')
    
    if request.method == 'POST':
        username_or_email = request.POST.get('username') or ''
        password = request.POST.get('password') or ''
        
        user = authenticate(request, username=username_or_email, password=password)
        if not user:
            try:
                user_obj = User.objects.get(email=username_or_email)
                user = authenticate(request, username=user_obj.username, password=password)
            except User.DoesNotExist:
                user = None
        
        if user and user.is_active:
            login(request, user)
            messages.success(request, 'Welcome back!')
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
            return redirect('dashboard')
        
        messages.error(request, 'Invalid credentials or inactive account.')
    
    return render(request, 'login.html', {'next': next_url})

@login_required
def dashboard_view(request):
    user = request.user
    context = {
        'user': user,
        'user_type': user.user_type,
    }
    return render(request, 'dashboard.html', context)

@login_required
def claims_management_view(request):
    if request.user.user_type not in ['manager', 'underwriter']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    if request.method == 'POST' and 'generate_report' in request.POST:
        messages.success(request, 'Report generation started. You will be notified when ready.')
    
    # Get claims for display
    claims = Claim.objects.select_related('policy__customer', 'policy__vehicle').order_by('-created_at')
    
    # Apply filters
    claim_id = request.GET.get('claim_id')
    vehicle_no = request.GET.get('vehicle_no')
    status_filter = request.GET.get('status')
    amount_range = request.GET.get('amount_range')
    
    if claim_id:
        claims = claims.filter(claim_id__icontains=claim_id)
    if vehicle_no:
        claims = claims.filter(policy__vehicle__vehicle_number__icontains=vehicle_no)
    if status_filter:
        claims = claims.filter(approval_status=status_filter)
    if amount_range:
        if amount_range == '0-25000':
            claims = claims.filter(estimated_amount__lte=25000)
        elif amount_range == '25001-50000':
            claims = claims.filter(estimated_amount__gte=25001, estimated_amount__lte=50000)
        elif amount_range == '50001-100000':
            claims = claims.filter(estimated_amount__gt=50000)
    
    context = {
        'user': request.user,
        'user_type': request.user.user_type,
        'claims': claims,
    }
    return render(request, 'claims_management.html', context)

@login_required
def chief_engineer_dashboard_view(request):
    if request.user.user_type not in ['chief_engineer', 'manager']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        messages.success(request, 'Report generated successfully!')
    
    context = {
        'user': request.user,
        'user_type': request.user.user_type,
    }
    return render(request, 'chief_engineer_dashboard.html', context)

def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    next_url = request.GET.get('next') or request.POST.get('next')
    
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        nic = request.POST.get('nic')
        address_no = request.POST.get('address_no')
        street = request.POST.get('street')
        town = request.POST.get('town')
        contact_no = request.POST.get('contact_no')
        email = request.POST.get('username_email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        if not all([first_name, last_name, nic, address_no, street, town, contact_no, email, password, confirm_password]):
            messages.error(request, 'Please fill in all required fields.')
            return render(request, 'register.html', {'next': next_url})
        
        if password != confirm_password:
            messages.error(request, "Passwords don't match.")
            return render(request, 'register.html', {'next': next_url})
        
        try:
            # Check if user already exists
            if User.objects.filter(email=email).exists():
                messages.error(request, 'A user with this email already exists.')
                return render(request, 'register.html', {'next': next_url})
            
            # Generate unique username
            username = email
            if User.objects.filter(username=username).exists():
                base = email.split('@')[0]
                suffix = 1
                while User.objects.filter(username=f"{base}{suffix}").exists():
                    suffix += 1
                username = f"{base}{suffix}"
            
            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                user_type='customer',
                first_name=first_name,
                last_name=last_name,
            )
            
            # Create customer profile (align with model fields)
            Customer.objects.create(
                user=user,
                address_no=address_no,
                street=street,
                town=town,
            )
            
            login(request, user)
            messages.success(request, 'Registration successful! Welcome to Drive Peak Insurance.')
            
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
            return redirect('dashboard')
            
        except Exception as exc:
            messages.error(request, f'Registration failed: {exc}')
            return render(request, 'register.html', {'next': next_url})
    
    context = {'next': next_url}
    return render(request, 'register.html', context)

# Additional API endpoints for AJAX calls
@extend_schema(tags=['claims'], summary='Update claim status')
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_claim_status(request, claim_id):
    """Update claim status via AJAX"""
    try:
        claim = Claim.objects.get(id=claim_id)
        action = request.data.get('action')
        
        if action == 'approve':
            claim.approval_status = 'approve'
            claim.status = 'approved'
        elif action == 'reject':
            claim.approval_status = 'reject'
            claim.status = 'rejected'
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})
        
        claim.processed_by = request.user
        claim.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Claim {action}d successfully',
            'status': claim.approval_status
        })
        
    except Claim.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Claim not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@extend_schema(tags=['dashboard'], summary='Get dashboard summary data')
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_data(request):
    """Get dashboard data for AJAX calls"""
    user = request.user
    
    if user.user_type == 'customer':
        try:
            customer = user.customer_profile
            data = {
                'active_policies': InsurancePolicy.objects.filter(customer=customer, status='active').count(),
                'total_claims': Claim.objects.filter(policy__customer=customer).count(),
                'pending_claims': Claim.objects.filter(policy__customer=customer, approval_status='pending').count(),
                'vehicles': Vehicle.objects.filter(customer=customer).count(),
            }
        except:
            data = {'active_policies': 0, 'total_claims': 0, 'pending_claims': 0, 'vehicles': 0}
    else:
        data = {
            'total_customers': Customer.objects.count(),
            'active_policies': InsurancePolicy.objects.filter(status='active').count(),
            'pending_claims': Claim.objects.filter(approval_status='pending').count(),
            'total_claims': Claim.objects.count(),
        }
    
    return JsonResponse(data)

@extend_schema(tags=['claims'], summary='Get claims data list')
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def claims_data(request):
    """Get claims data for dashboard and management interface.
    - Customers: only their own claims
    - Staff (underwriter, manager): all claims
    """
    user = request.user
    if getattr(user, 'user_type', None) == 'customer':
        try:
            customer = user.customer_profile
            claims = Claim.objects.filter(policy__customer=customer).order_by('-created_at')
        except Exception:
            claims = Claim.objects.none()
    else:
        claims = Claim.objects.all().order_by('-created_at')
    
    # Apply filters
    claim_id = request.GET.get('claim_id')
    vehicle_no = request.GET.get('vehicle_no')
    status_filter = request.GET.get('status')
    
    if claim_id:
        claims = claims.filter(claim_id__icontains=claim_id)
    if vehicle_no:
        claims = claims.filter(policy__vehicle__vehicle_number__icontains=vehicle_no)
    if status_filter:
        claims = claims.filter(approval_status=status_filter)
    
    claims_data = []
    for claim in claims:
        customer = getattr(claim.policy, 'customer', None)
        vehicle = getattr(claim.policy, 'vehicle', None)
        customer_name = customer.user.get_full_name() if customer else 'N/A'
        claims_data.append({
            'id': claim.id,
            'claim_id': claim.claim_id,
            'vehicle_number': getattr(vehicle, 'vehicle_number', 'N/A'),
            'estimated_amount': str(claim.estimated_amount),
            'approval_status': claim.approval_status,
            'status': claim.status,
            'customer_name': customer_name or 'N/A',
            'created_at': claim.created_at.strftime('%Y-%m-%d %H:%M'),
        })
    
    return JsonResponse({'claims': claims_data})

# Template context processor function (add to settings.py)
def user_context(request):
    """Add user context to all templates"""
    context = {}
    if request.user.is_authenticated:
        context['user'] = request.user
        context['user_type'] = request.user.user_type
        context['is_customer'] = request.user.user_type == 'customer'
        context['is_staff'] = request.user.is_staff or request.user.user_type in ['manager', 'underwriter']
        
        # Add customer profile if exists
        if request.user.user_type == 'customer':
            try:
                context['customer_profile'] = request.user.customer_profile
            except:
                context['customer_profile'] = None
    
    return context

# Auth API - Register
@extend_schema(
    request=RegisterSerializer,
    responses={
        201: inline_serializer(
            name='RegisterResponse',
            fields={
                'user': UserSerializer(),
                'access': rf_serializers.CharField(),
                'refresh': rf_serializers.CharField(),
            }
        ),
        400: inline_serializer(name='RegisterError', fields={'error': rf_serializers.CharField()}),
    },
    summary='Register a new customer and receive JWT tokens',
    tags=['auth']
)
@api_view(['POST'])
@permission_classes([AllowAny])
def register_api(request):
    """Register a new customer user and start 2FA verification (no tokens yet)."""
    serializer = RegisterSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    try:
        from django.utils import timezone
        import random
        with transaction.atomic():
            user = serializer.save()
            # generate 6-digit code, valid for 10 minutes
            code = f"{random.randint(0, 999999):06d}"
            TwoFactorCode.objects.filter(user=user, is_used=False).update(is_used=True)
            TwoFactorCode.objects.create(
                user=user,
                code=code,
                expires_at=timezone.now() + timedelta(minutes=10)
            )
            # send OTP via email/SMS
            try:
                send_otp(user, code)
            except Exception:
                # don't fail registration if messaging fails; client can resend
                pass
            # TODO: send code via email/SMS provider. For development, return a hint.
            return Response({
                'message': 'Registration successful. Please verify the OTP sent to your email/phone to activate your account.',
                'user': UserSerializer(user).data,
                'otp_hint': code if settings.DEBUG else None,
            }, status=status.HTTP_201_CREATED)
    except Exception as exc:
        return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

@extend_schema(
    request=inline_serializer(name='VerifyOtpRequest', fields={'email': rf_serializers.EmailField(), 'code': rf_serializers.CharField(max_length=6)}),
    responses={200: inline_serializer(name='VerifyOtpResponse', fields={'message': rf_serializers.CharField()})},
    tags=['auth'], summary='Verify registration OTP'
)
@api_view(['POST'])
@permission_classes([AllowAny])
def verify_registration_otp(request):
    from django.utils import timezone
    email = (request.data.get('email') or '').strip().lower()
    code = (request.data.get('code') or '').strip()
    if not email or not code:
        return Response({'error': 'email and code are required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    tf = TwoFactorCode.objects.filter(user=user, is_used=False).order_by('-created_at').first()
    if not tf:
        return Response({'error': 'No active OTP. Please request a new code.'}, status=status.HTTP_400_BAD_REQUEST)
    if tf.expires_at < timezone.now():
        return Response({'error': 'OTP expired. Please request a new code.'}, status=status.HTTP_400_BAD_REQUEST)
    tf.attempts += 1
    if tf.attempts > 5:
        tf.is_used = True
        tf.save(update_fields=['attempts', 'is_used'])
        return Response({'error': 'Too many attempts. OTP invalidated. Request a new code.'}, status=status.HTTP_400_BAD_REQUEST)
    if tf.code != code:
        tf.save(update_fields=['attempts'])
        return Response({'error': 'Invalid code.'}, status=status.HTTP_400_BAD_REQUEST)

    # success
    tf.is_used = True
    tf.save(update_fields=['is_used'])
    if not user.is_active:
        user.is_active = True
        user.save(update_fields=['is_active'])
    return Response({'message': 'Verification successful. You can now log in.'})

@extend_schema(
    request=inline_serializer(name='ResendOtpRequest', fields={'email': rf_serializers.EmailField()}),
    responses={200: inline_serializer(name='ResendOtpResponse', fields={'message': rf_serializers.CharField(), 'otp_hint': rf_serializers.CharField(required=False)})},
    tags=['auth'], summary='Resend registration OTP'
)
@api_view(['POST'])
@permission_classes([AllowAny])
def resend_registration_otp(request):
    from django.utils import timezone
    import random
    email = (request.data.get('email') or '').strip().lower()
    if not email:
        return Response({'error': 'email is required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    if user.is_active:
        return Response({'error': 'User already verified.'}, status=status.HTTP_400_BAD_REQUEST)
    # Invalidate previous active codes
    TwoFactorCode.objects.filter(user=user, is_used=False).update(is_used=True)
    code = f"{random.randint(0, 999999):06d}"
    TwoFactorCode.objects.create(user=user, code=code, expires_at=timezone.now() + timedelta(minutes=10))
    # send OTP via email/SMS
    try:
        send_otp(user, code)
    except Exception:
        pass
    return Response({'message': 'OTP resent.', 'otp_hint': code if settings.DEBUG else None})

# --- JWT obtain via email + password ---
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.permissions import AllowAny


class EmailTokenObtainPairSerializer(TokenObtainPairSerializer):
    """Allow authentication with email + password in addition to username."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Make username optional so missing username doesn't 400 before validate
        if 'username' in self.fields:
            self.fields['username'].required = False
        # Add optional email field for documentation; not strictly needed for parsing
        from rest_framework import serializers as rf_serializers
        self.fields['email'] = rf_serializers.EmailField(required=False, write_only=True)

    def validate(self, attrs):
        # Use initial_data to access 'email' even if it's not a declared field
        data = dict(self.initial_data)
        email = (data.get('email') or '').strip().lower()
        username = data.get('username') or attrs.get('username')
        password = data.get('password') or attrs.get('password')

        # Map email to username if provided and username missing
        if email and not username:
            User = get_user_model()
            try:
                user = User.objects.get(email=email)
                username = user.get_username()
            except User.DoesNotExist:
                # Fallback to email to force auth failure with standard error (no user enumeration)
                username = email

        # Delegate to parent with normalized attrs
        normalized = {'username': username, 'password': password}
        return super().validate(normalized)


class EmailTokenObtainPairView(TokenObtainPairView):
    serializer_class = EmailTokenObtainPairSerializer
    permission_classes = [AllowAny]


# --- Password Reset ---
@api_view(["POST"])
@permission_classes([AllowAny])
def password_reset_request(request):
    """Accepts { email } and sends a reset link if user exists. Responds 200 regardless to avoid user enumeration."""
    email = (request.data.get("email") or "").strip().lower()
    if not email:
        return Response({"error": "Email is required"}, status=status.HTTP_400_BAD_REQUEST)
    User = get_user_model()
    try:
        user = User.objects.get(email=email)
        token = PasswordResetTokenGenerator().make_token(user)
        uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
        reset_url = f"{settings.FRONTEND_BASE_URL}/reset-password/{uidb64}/{token}"
        subject = "Password Reset Request"
        message = f"Click the link to reset your password: {reset_url}\nIf you did not request this, you can ignore this email."
        try:
            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email], fail_silently=True)
        except Exception:
            pass
    except User.DoesNotExist:
        # Do not reveal existence
        pass
    return Response({"message": "If the email exists, a reset link has been sent."})


@api_view(["POST"])
@permission_classes([AllowAny])
def password_reset_confirm(request):
    """Accepts { uidb64, token, new_password } and resets the password if token is valid."""
    uidb64 = request.data.get("uidb64")
    token = request.data.get("token")
    new_password = request.data.get("new_password")
    if not uidb64 or not token or not new_password:
        return Response({"error": "uidb64, token and new_password are required"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        User = get_user_model()
        user = User.objects.get(pk=uid)
    except Exception:
        return Response({"error": "Invalid link"}, status=status.HTTP_400_BAD_REQUEST)

    if not PasswordResetTokenGenerator().check_token(user, token):
        return Response({"error": "Invalid or expired token"}, status=status.HTTP_400_BAD_REQUEST)

    # Validate password with Django validators
    try:
        validate_password(new_password, user=user)
    except Exception as exc:
        return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    user.set_password(new_password)
    user.save()
    return Response({"message": "Password has been reset successfully."})